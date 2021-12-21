from openpyxl import Workbook, load_workbook
import PySimpleGUI as sg

layout = [
		[sg.Text('Digite o Caminho da Pasta com o Arquivo Excel:'), sg.Input(size=(45, 1), key='CaminhoArquivo')],
        [sg.Text('Digite o Nome do Arquivo Excel: ' + 23*' '), sg.Input(size=(45, 1), key='NomeArquivo')],
		[sg.Button('Ok'), sg.Button('Cancel')]
	]


window = sg.Window('Geração de Arquivo de VT', layout)

while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Cancel':
        break
    
    path = values['CaminhoArquivo']
    fname = values['NomeArquivo'] + '.xlsx'
    full_path = path + '\\' + fname

    wb = load_workbook(full_path, data_only=True)


    planilhas = {}


    for p in range(len(wb.sheetnames)):
        planilhas[wb.sheetnames[p]] = {'matriculas':'', 'valores':''}


    for sheet in wb.sheetnames:
        ws = wb[sheet]

        colB = ws['B']
        matriculas = []
        for m in colB:
            if (m.value != None) and (type(m.value) != str):
                matriculas.append(str(m.value))
        planilhas[sheet]['matriculas'] = matriculas

        colI = ws['I']
        valores = []
        for i in colI:
            if (i.value != None) and (type(i.value) != str):
                valores.append(format(i.value, '.2f'))
        for v in range(len(valores)):
            valores[v] = valores[v].replace('.', '')    
        planilhas[sheet]['valores'] = valores


        linha_01 = "0000101PEDIDO01.0003686998000118"
        mat = []
        totais = [] 
        sequencia = []
        cargas = []
        total_cargas = []
        trailler = []
        
        mat = planilhas[sheet]['matriculas']
        totais = planilhas[sheet]['valores']
        
        total_cargas.append(planilhas[sheet]['valores'][-1])
        totais.remove(totais[-1])

        for i in range(len(mat)):   
            sequencia.append(i + 2)


        for j in range(len(sequencia)): 
            sequencia[j] = str(sequencia[j]).zfill(5) + '02'   

            
        for j in range(len(mat)):    
            if len(mat[j]) < 15:
                mat[j] = mat[j] + (15 - len(mat[j])) * ' '


        for i in range(len(totais)):
            if len(totais[i]) < 8:
                totais[i] = (8 - len(totais[i])) * '0' + totais[i]

        total_linhas = len(mat) + 2
        total_linhas = str(total_linhas)

        trailler.append(total_linhas)

        for j in range(len(trailler)):  
            if len(trailler[j]) < 5:
                trailler[j] = (5 - len(trailler[j])) * '0' + trailler[j]

        trailler[0] = trailler[0] + '99'


        if len(total_cargas[0]) < 10:
            total_cargas[0] = (10 - len(total_cargas[0])) * '0' + total_cargas[0]


        trailler[0] = trailler[0] + total_cargas[0]

        fname_recarga = sheet.lower() + '.txt'

        with open(path + '\\' + fname_recarga, 'w') as f:
            print(linha_01, file=f)
            for j in range(len(totais)):
                print(sequencia[j], mat[j], totais[j], sep='', file=f)
            print(trailler[0], file=f)        
    
window.close()